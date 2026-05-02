"""
NACH Monthly Sheet Generator
------------------------------
Pulls every active loanshare with an EMI due in a given month from Kosh DB,
joins loan_id and all three NACH providers (Setu / LotsuPay / Digio),
flags each mandate as Active / Inactive / No NACH, and exports to Excel.

Usage:
  python nach_monthly_sheet.py              # current month
  python nach_monthly_sheet.py 2026-04      # specific month (YYYY-MM)

Output (Desktop):
  nach_sheet_YYYY_MM.xlsx   -- full row-level data
  nach_chart_YYYY_MM.png    -- summary bar chart
"""

import sys
import os
import requests
import calendar
from datetime import date, datetime

# ── Config ────────────────────────────────────────────────────────────────────
BASE_URL  = "https://superset.bkosh.com"
USERNAME  = "ayushd"
PASSWORD  = "Getkosh101"
KOSH_DB   = 1                                       # Kosh DB (production)
DESKTOP   = os.path.join(os.path.expanduser("~"), "Desktop")
PAGE_SIZE = 9_000   # stay under Superset's 10k row cap per request

# ── Auth ──────────────────────────────────────────────────────────────────────
def get_auth():
    session = requests.Session()
    resp = session.post(
        f"{BASE_URL}/api/v1/security/login",
        json={"username": USERNAME, "password": PASSWORD, "provider": "db", "refresh": True},
    )
    resp.raise_for_status()
    token = resp.json()["access_token"]
    csrf  = session.get(
        f"{BASE_URL}/api/v1/security/csrf_token/",
        headers={"Authorization": f"Bearer {token}"},
    ).json()["result"]
    return session, token, csrf


def run_query(session, token, csrf, sql, db_id=KOSH_DB):
    resp = session.post(
        f"{BASE_URL}/api/v1/sqllab/execute/",
        json={"database_id": db_id, "sql": sql},
        headers={
            "Authorization": f"Bearer {token}",
            "X-CSRFToken": csrf,
            "Content-Type": "application/json",
            "Referer": f"{BASE_URL}/sqllab/",
        },
    )
    if resp.status_code != 200:
        raise RuntimeError(f"Query failed [{resp.status_code}]: {resp.text[:500]}")
    result = resp.json()
    if "data" not in result:
        raise RuntimeError(f"Unexpected response: {str(result)[:400]}")
    return result["data"]


def run_paged(session, token, csrf, base_sql: str, db_id=KOSH_DB):
    """Run a query in pages of PAGE_SIZE rows and concatenate all results."""
    all_rows = []
    offset   = 0
    while True:
        paged_sql = f"{base_sql}\nLIMIT {PAGE_SIZE} OFFSET {offset}"
        page = run_query(session, token, csrf, paged_sql, db_id)
        all_rows.extend(page)
        print(f"  ... fetched {len(all_rows):,} rows", end="\r")
        if len(page) < PAGE_SIZE:
            break
        offset += PAGE_SIZE
    print()
    return all_rows


# ── Query ─────────────────────────────────────────────────────────────────────
def fetch_rows(session, token, csrf, month_start: str, month_end: str):
    """One row per installment, joined with loan_id and best NACH mandate."""
    sql = f"""
WITH digio_best AS (
    -- One Digio mandate per loanshare: prefer auth_success, then latest
    SELECT DISTINCT ON (loanshare_id)
        loanshare_id,
        mandate_id   AS digio_mandate_id,
        status       AS digio_status
    FROM nach_digiomandate
    ORDER BY loanshare_id,
        CASE WHEN status = 'auth_success' THEN 0 ELSE 1 END,
        created_at DESC
)
SELECT
    ll.id                                               AS loan_id,
    ls.id                                               AS loanshare_id,
    i.id                                                AS installment_id,
    i.number                                            AS emi_number,
    i.due_date                                          AS emi_due_date,
    i.amount                                            AS emi_amount,

    -- NACH provider
    CASE
        WHEN ls.upi_reccur_mandate_id IS NOT NULL THEN 'Setu UPI'
        WHEN ls.mandate_id            IS NOT NULL THEN 'LotsuPay'
        WHEN dg.loanshare_id          IS NOT NULL THEN 'Digio'
        ELSE 'No NACH'
    END                                                 AS nach_provider,

    -- Mandate ID
    COALESCE(
        ls.upi_reccur_mandate_id,
        ls.mandate_id,
        dg.digio_mandate_id::text
    )                                                   AS mandate_id,

    -- Raw status from respective table
    COALESCE(
        sm.status,
        nm.status,
        dg.digio_status
    )                                                   AS mandate_status,

    -- Active flag
    CASE
        WHEN sm.status    = 'live'         THEN 'Active'
        WHEN nm.status    = 'active'       THEN 'Active'
        WHEN dg.digio_status = 'auth_success' THEN 'Active'
        WHEN ls.upi_reccur_mandate_id IS NOT NULL
          OR ls.mandate_id IS NOT NULL
          OR dg.loanshare_id IS NOT NULL   THEN 'Inactive'
        ELSE 'No NACH'
    END                                                 AS nach_active

FROM loan_installment i

JOIN loan_loanshare   ls ON i.loanshare_id  = ls.id
JOIN loan_loanapp     la ON ls.loanapp_id   = la.id
JOIN loan_loan        ll ON la.loan_id      = ll.id

LEFT JOIN nach_setumandate sm ON ls.upi_reccur_mandate_id = sm.mandate_id
LEFT JOIN nach_mandate     nm ON ls.mandate_id            = nm.mandate_id
LEFT JOIN digio_best       dg ON ls.id                    = dg.loanshare_id

WHERE i.due_date >= '{month_start}'
  AND i.due_date  < '{month_end}'
  AND i.status    = 'opened'
  AND ls.user_status = 'loan_disbursed'

ORDER BY ll.id, ls.id, i.due_date
"""
    return run_paged(session, token, csrf, sql)


# ── Summary ───────────────────────────────────────────────────────────────────
def build_summary(rows):
    from collections import defaultdict
    summary = defaultdict(lambda: {"Active": 0, "Inactive": 0, "No NACH": 0,
                                   "emi_active": 0, "emi_inactive": 0, "emi_no_nach": 0})
    for r in rows:
        prov   = r["nach_provider"]
        flag   = r["nach_active"]
        amount = r["emi_amount"] or 0
        summary[prov][flag] += 1
        if   flag == "Active":   summary[prov]["emi_active"]   += amount
        elif flag == "Inactive": summary[prov]["emi_inactive"] += amount
        else:                    summary[prov]["emi_no_nach"]  += amount
    return dict(summary)


# ── Print report ──────────────────────────────────────────────────────────────
def print_report(summary, month_label: str):
    print(f"\n{'='*80}")
    print(f"  NACH Sheet Report - {month_label}")
    print(f"{'='*80}")
    print(f"{'Provider':<12} {'Active':>8} {'Inactive':>10} {'No NACH':>9} {'Total':>8}  "
          f"{'Active EMI (Rs)':>16} {'Total EMI (Rs)':>16}")
    print(f"{'-'*80}")

    g = {"Active": 0, "Inactive": 0, "No NACH": 0, "emi": 0}
    for prov, s in sorted(summary.items()):
        total     = s["Active"] + s["Inactive"] + s["No NACH"]
        total_emi = s["emi_active"] + s["emi_inactive"] + s["emi_no_nach"]
        pct = f"({s['Active']/total*100:.1f}%)" if total else ""
        print(f"{prov:<12} {s['Active']:>6,} {pct:>4}  {s['Inactive']:>8,}  "
              f"{s['No NACH']:>7,}  {total:>6,}  "
              f"{s['emi_active']:>16,.0f} {total_emi:>16,.0f}")
        g["Active"]   += s["Active"]
        g["Inactive"] += s["Inactive"]
        g["No NACH"]  += s["No NACH"]
        g["emi"]      += total_emi

    grand_total = g["Active"] + g["Inactive"] + g["No NACH"]
    g_pct = f"({g['Active']/grand_total*100:.1f}%)" if grand_total else ""
    print(f"{'-'*80}")
    print(f"{'TOTAL':<12} {g['Active']:>6,} {g_pct:>4}  {g['Inactive']:>8,}  "
          f"{g['No NACH']:>7,}  {grand_total:>6,}  {'':>16} {g['emi']:>16,.0f}")
    print(f"\n  Total EMI Demand : Rs {g['emi']:,.0f}")
    print(f"{'='*80}\n")


# ── Excel export ──────────────────────────────────────────────────────────────
def save_excel(rows, summary, month_label: str, out_path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Sheet 1: Detail ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Detail"

    COLS = ["loan_id", "loanshare_id", "installment_id", "emi_number",
            "emi_due_date", "emi_amount", "nach_provider",
            "mandate_id", "mandate_status", "nach_active"]
    HEADERS = ["Loan ID", "Loanshare ID", "Installment ID", "EMI #",
               "Due Date", "EMI Amount (Rs)", "NACH Provider",
               "Mandate ID", "Mandate Status", "NACH Active"]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Color fills for nach_active
    fill_active   = PatternFill("solid", fgColor="C6EFCE")
    fill_inactive = PatternFill("solid", fgColor="FFC7CE")
    fill_no_nach  = PatternFill("solid", fgColor="FFEB9C")

    for row_idx, r in enumerate(rows, 2):
        flag = r.get("nach_active", "")
        row_fill = (fill_active   if flag == "Active"
                    else fill_inactive if flag == "Inactive"
                    else fill_no_nach)
        for col_idx, key in enumerate(COLS, 1):
            val  = r.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if key == "nach_active":
                cell.fill = row_fill
                cell.font = Font(bold=True)
            elif key == "emi_amount":
                cell.number_format = '#,##0'

    # Auto-fit columns
    for col_idx, _ in enumerate(COLS, 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, min(len(rows) + 2, 500))
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Summary ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.title = "Summary"

    s_headers = ["Provider", "Active", "Inactive", "No NACH", "Total",
                 "Active EMI (Rs)", "Inactive EMI (Rs)", "No NACH EMI (Rs)", "Total EMI (Rs)", "Activation %"]
    for ci, h in enumerate(s_headers, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row = 2
    g = {"Active": 0, "Inactive": 0, "No NACH": 0,
         "emi_active": 0, "emi_inactive": 0, "emi_no_nach": 0}
    for prov, s in sorted(summary.items()):
        total     = s["Active"] + s["Inactive"] + s["No NACH"]
        total_emi = s["emi_active"] + s["emi_inactive"] + s["emi_no_nach"]
        pct = round(s["Active"] / total * 100, 1) if total else 0
        vals = [prov, s["Active"], s["Inactive"], s["No NACH"], total,
                s["emi_active"], s["emi_inactive"], s["emi_no_nach"], total_emi, pct]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=ci, value=v)
            cell.border = border
            if ci in (2, 3, 4, 5):
                cell.number_format = '#,##0'
            if ci in (6, 7, 8, 9):
                cell.number_format = '#,##0'
            if ci == 10:
                cell.number_format = '0.0"%"'
        for k in g: g[k] += s[k]
        row += 1

    # Grand total row
    grand_total = g["Active"] + g["Inactive"] + g["No NACH"]
    grand_emi   = g["emi_active"] + g["emi_inactive"] + g["emi_no_nach"]
    grand_pct   = round(g["Active"] / grand_total * 100, 1) if grand_total else 0
    total_fill  = PatternFill("solid", fgColor="BDD7EE")
    total_font  = Font(bold=True)
    for ci, v in enumerate(["TOTAL", g["Active"], g["Inactive"], g["No NACH"], grand_total,
                             g["emi_active"], g["emi_inactive"], g["emi_no_nach"], grand_emi, grand_pct], 1):
        cell = ws2.cell(row=row, column=ci, value=v)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = border
        if ci in (2, 3, 4, 5, 6, 7, 8, 9):
            cell.number_format = '#,##0'
        if ci == 10:
            cell.number_format = '0.0"%"'

    for ci in range(1, len(s_headers) + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 20

    ws2.freeze_panes = "A2"

    wb.save(out_path)
    print(f"Excel saved -> {out_path}")


# ── Chart ─────────────────────────────────────────────────────────────────────
def save_chart(summary, month_label: str, out_path: str):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import numpy as np

    providers = sorted(summary.keys())
    active    = [summary[p]["Active"]   for p in providers]
    inactive  = [summary[p]["Inactive"] for p in providers]
    no_nach   = [summary[p]["No NACH"]  for p in providers]

    x     = np.arange(len(providers))
    width = 0.25

    fig, ax = plt.subplots(figsize=(10, 6))
    b1 = ax.bar(x - width, active,   width, label="Active",   color="#2ecc71")
    b2 = ax.bar(x,         inactive, width, label="Inactive",  color="#e74c3c")
    b3 = ax.bar(x + width, no_nach,  width, label="No NACH",   color="#95a5a6")

    def label_bars(bars):
        for bar in bars:
            h = bar.get_height()
            if h > 0:
                ax.text(bar.get_x() + bar.get_width() / 2, h + 50,
                        f"{int(h):,}", ha="center", va="bottom", fontsize=8)

    for b in (b1, b2, b3):
        label_bars(b)

    ax.set_title(f"NACH Mandate Status - {month_label}", fontsize=14, fontweight="bold")
    ax.set_xlabel("NACH Provider")
    ax.set_ylabel("Installments")
    ax.set_xticks(x)
    ax.set_xticklabels(providers)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{int(v):,}"))
    ax.legend()
    ax.grid(axis="y", linestyle="--", alpha=0.4)
    plt.tight_layout()
    plt.savefig(out_path, dpi=150)
    plt.close()
    print(f"Chart saved  -> {out_path}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    target = (
        datetime.strptime(sys.argv[1], "%Y-%m") if len(sys.argv) > 1
        else datetime(date.today().year, date.today().month, 1)
    )
    year, month = target.year, target.month
    next_m      = date(year + (month // 12), (month % 12) + 1, 1)
    month_start = f"{year}-{month:02d}-01"
    month_end   = next_m.strftime("%Y-%m-%d")
    month_label = target.strftime("%B %Y")
    slug        = f"{year}_{month:02d}"

    print(f"Authenticating to {BASE_URL} ...")
    session, token, csrf = get_auth()
    print(f"Auth OK | Kosh DB (ID: {KOSH_DB}) | Month: {month_label}")

    print("Fetching installment + NACH data ...")
    rows = fetch_rows(session, token, csrf, month_start, month_end)
    print(f"Fetched {len(rows):,} rows")

    summary = build_summary(rows)
    print_report(summary, month_label)

    xlsx_path  = os.path.join(DESKTOP, f"nach_sheet_{slug}.xlsx")
    chart_path = os.path.join(DESKTOP, f"nach_chart_{slug}.png")

    print("Writing Excel file ...")
    save_excel(rows, summary, month_label, xlsx_path)
    save_chart(summary, month_label, chart_path)


if __name__ == "__main__":
    main()
