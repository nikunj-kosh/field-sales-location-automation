"""
Setu NACH Monthly Sheet
=======================
Answers: "For every active loan with an EMI due this month,
          does it have a working Setu UPI NACH mandate?"

Data flow
---------
loan_installment  (EMI due this month, status='opened')
  → loan_loanshare        (user_status='loan_disbursed')
    → loan_loanapp / loan_loan   (to get Loan ID)
    → nach_setumandate           (the UPI AutoPay mandate)
      → Payment_setunach         (one row per debit attempt)
           latest row's enabled=True  → Active
           latest row's enabled=False → Disabled
           no row at all              → Not Yet Attempted

Usage
-----
  python nach_setu_sheet.py              # current month
  python nach_setu_sheet.py 2026-04      # any month (YYYY-MM)

Output (Desktop)
----------------
  setu_nach_sheet_YYYY_MM.xlsx  —  Detail tab + Summary tab
  setu_nach_chart_YYYY_MM.png   —  bar + pie chart
"""

import sys
import os
import requests
from datetime import date, datetime

# ── Config ────────────────────────────────────────────────────────────────────
BASE_URL  = os.environ.get("SUPERSET_BASE_URL", "https://superset.bkosh.com")
USERNAME  = os.environ.get("SUPERSET_UN", "ayushd")
PASSWORD  = os.environ.get("SUPERSET_PASS", "Getkosh101")
KOSH_DB   = int(os.environ.get("SUPERSET_KOSH_DB", "1"))
OUTPUT_DIR = os.environ.get(
    "NACH_OUTPUT_DIR",
    os.path.join(os.path.expanduser("~"), "Desktop"),
)
PAGE_SIZE = 9_000   # stay under Superset's 10k-row cap per request

# ── Auth ──────────────────────────────────────────────────────────────────────
def get_auth():
    session = requests.Session()
    resp = session.post(
        f"{BASE_URL}/api/v1/security/login",
        json={"username": USERNAME, "password": PASSWORD,
              "provider": "db", "refresh": True},
    )
    resp.raise_for_status()
    token = resp.json()["access_token"]
    csrf  = session.get(
        f"{BASE_URL}/api/v1/security/csrf_token/",
        headers={"Authorization": f"Bearer {token}"},
    ).json()["result"]
    return session, token, csrf


def run_query(session, token, csrf, sql):
    resp = session.post(
        f"{BASE_URL}/api/v1/sqllab/execute/",
        json={"database_id": KOSH_DB, "sql": sql},
        headers={
            "Authorization": f"Bearer {token}",
            "X-CSRFToken": csrf,
            "Content-Type": "application/json",
            "Referer": f"{BASE_URL}/sqllab/",
        },
    )
    if resp.status_code != 200:
        raise RuntimeError(f"Query failed [{resp.status_code}]: {resp.text[:500]}")
    result = resp.json()
    if "data" not in result:
        raise RuntimeError(f"Unexpected response: {str(result)[:400]}")
    return result["data"]


def run_paged(session, token, csrf, base_sql: str):
    """Execute a query in pages of PAGE_SIZE rows to bypass Superset's row cap."""
    all_rows, offset = [], 0
    while True:
        page = run_query(session, token, csrf,
                         f"{base_sql}\nLIMIT {PAGE_SIZE} OFFSET {offset}")
        all_rows.extend(page)
        print(f"  ... fetched {len(all_rows):,} rows", end="\r")
        if len(page) < PAGE_SIZE:
            break
        offset += PAGE_SIZE
    print()
    return all_rows


# ── Query ─────────────────────────────────────────────────────────────────────
def fetch_rows(session, token, csrf, month_start: str, month_end: str):
    """
    Returns one row per installment for active loanshares in the target month,
    joined with Setu mandate and the latest Payment_setunach record.

    Key output columns:
      nach_enabled  — True/False/None  (Payment_setunach.enabled of latest attempt)
      nach_active   — 'Active' / 'Disabled' / 'Not Yet Attempted'
    """
    sql = f"""
WITH max_seq AS (
    -- Find the highest sequence_number per mandate (= the most recent attempt).
    -- Using GROUP BY + JOIN instead of DISTINCT ON to avoid temp-disk pressure.
    SELECT mandate_id, MAX(sequence_number) AS max_seq
    FROM "Payment_setunach"
    GROUP BY mandate_id
),
latest_payment AS (
    -- Pull the full row for that latest sequence
    SELECT
        ps.mandate_id,
        ps.sequence_number    AS latest_seq,
        ps.execution_time     AS latest_exec_time,
        ps.status             AS latest_payment_status,
        ps.enabled            AS nach_enabled
    FROM "Payment_setunach" ps
    JOIN max_seq ms
      ON ps.mandate_id       = ms.mandate_id
     AND ps.sequence_number  = ms.max_seq
)
SELECT
    -- Loan / loanshare identifiers
    ll.id                       AS loan_id,
    ls.id                       AS loanshare_id,

    -- Installment details
    i.id                        AS installment_id,
    i.number                    AS emi_number,
    i.due_date                  AS emi_due_date,
    i.amount                    AS emi_amount,

    -- Setu mandate details
    sm.mandate_id               AS setu_mandate_id,
    sm.status                   AS mandate_status,
    sm.start_date               AS mandate_start_date,
    sm.end_date                 AS mandate_end_date,

    -- Latest payment attempt
    lp.latest_seq,
    lp.latest_exec_time,
    lp.latest_payment_status,
    lp.nach_enabled,

    -- Human-readable active flag
    CASE
        WHEN lp.mandate_id IS NULL   THEN 'Not Yet Attempted'
        WHEN lp.nach_enabled = true  THEN 'Active'
        WHEN lp.nach_enabled = false THEN 'Disabled'
    END                         AS nach_active

FROM loan_installment i

JOIN loan_loanshare   ls ON i.loanshare_id  = ls.id
JOIN loan_loanapp     la ON ls.loanapp_id   = la.id
JOIN loan_loan        ll ON la.loan_id      = ll.id

-- Only loanshares that have a Setu UPI mandate
JOIN nach_setumandate sm ON ls.upi_reccur_mandate_id = sm.mandate_id

-- Latest payment attempt (NULL if no attempt yet)
LEFT JOIN latest_payment lp ON sm.mandate_id = lp.mandate_id

WHERE i.due_date    >= '{month_start}'
  AND i.due_date     < '{month_end}'
  AND i.status       = 'opened'
  AND ls.user_status = 'loan_disbursed'

ORDER BY ll.id, ls.id, i.due_date
"""
    return run_paged(session, token, csrf, sql)


# ── Build summaries from fetched rows ─────────────────────────────────────────
def build_summary(rows):
    """
    Returns three breakdown dicts, each keyed by status string:
      active_s   — by nach_active         (Active / Disabled / Not Yet Attempted)
      mandate_s  — by mandate_status      (live / pending / initiated / revoked …)
      payment_s  — by latest_payment_status (success / failed / None …)
    Each value: {"count": int, "emi": int}
    """
    from collections import defaultdict

    active_s  = defaultdict(lambda: {"count": 0, "emi": 0})
    mandate_s = defaultdict(lambda: {"count": 0, "emi": 0})
    payment_s = defaultdict(lambda: {"count": 0, "emi": 0})

    for r in rows:
        amt = r["emi_amount"] or 0
        active_s [r["nach_active"]            or "Not Yet Attempted"]["count"] += 1
        active_s [r["nach_active"]            or "Not Yet Attempted"]["emi"]   += amt
        mandate_s[r["mandate_status"]         or "unknown"]["count"]           += 1
        mandate_s[r["mandate_status"]         or "unknown"]["emi"]             += amt
        payment_s[r["latest_payment_status"]  or "no attempt yet"]["count"]    += 1
        payment_s[r["latest_payment_status"]  or "no attempt yet"]["emi"]      += amt

    return dict(active_s), dict(mandate_s), dict(payment_s)


# ── Console report ────────────────────────────────────────────────────────────
def print_report(summaries, month_label: str):
    active_s, _, _ = summaries
    total = sum(v["count"] for v in active_s.values())
    total_emi = sum(v["emi"] for v in active_s.values())
    print(f"\n{'='*68}")
    print(f"  Setu NACH - {month_label}")
    print(f"{'='*68}")
    print(f"{'Status':<22} {'Count':>8}  {'%':>6}  {'EMI Demand (Rs)':>16}")
    print(f"{'-'*68}")
    for flag in ["Active", "Disabled", "Not Yet Attempted"]:
        v = active_s.get(flag, {"count": 0, "emi": 0})
        pct = v["count"] / total * 100 if total else 0
        print(f"{flag:<22} {v['count']:>8,}  {pct:>5.1f}%  {v['emi']:>16,.0f}")
    print(f"{'-'*68}")
    print(f"{'TOTAL':<22} {total:>8,}  {'100.0%':>6}  {total_emi:>16,.0f}")
    print(f"\n  Total EMI Demand : Rs {total_emi:,.0f}")
    print(f"{'='*68}\n")


# ── Excel export ──────────────────────────────────────────────────────────────
def save_excel(rows, summaries, month_label: str, out_path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    active_s, mandate_s, payment_s = summaries
    wb = openpyxl.Workbook()

    # ── Shared styles ─────────────────────────────────────────────────────────
    hdr_fill    = PatternFill("solid", fgColor="1F4E79")
    hdr_font    = Font(color="FFFFFF", bold=True)
    sec_fill    = PatternFill("solid", fgColor="2E4057")
    sec_font    = Font(color="FFFFFF", bold=True, size=11)
    kpi_fill    = PatternFill("solid", fgColor="E8F4FD")
    kpi_font    = Font(bold=True, size=10)
    tot_fill    = PatternFill("solid", fgColor="BDD7EE")
    fill_green  = PatternFill("solid", fgColor="C6EFCE")
    fill_red    = PatternFill("solid", fgColor="FFC7CE")
    fill_yellow = PatternFill("solid", fgColor="FFEB9C")
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    NUM, PCT    = '#,##0', '0.0"%"'

    # ── Detail tab ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Detail"

    COLS = [
        "loan_id", "loanshare_id", "installment_id", "emi_number",
        "emi_due_date", "emi_amount",
        "setu_mandate_id", "mandate_status", "mandate_start_date", "mandate_end_date",
        "latest_seq", "latest_exec_time", "latest_payment_status",
        "nach_enabled", "nach_active",
    ]
    HEADERS = [
        "Loan ID", "Loanshare ID", "Installment ID", "EMI #",
        "Due Date", "EMI Amount (Rs)",
        "Setu Mandate ID", "Mandate Status", "Mandate Start", "Mandate End",
        "Last Seq #", "Last Exec Time", "Last Payment Status",
        "NACH Enabled", "NACH Active",
    ]

    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill, c.font = hdr_fill, hdr_font
        c.alignment = Alignment(horizontal="center")
        c.border = border

    for ri, r in enumerate(rows, 2):
        flag      = r.get("nach_active", "")
        row_fill  = (fill_green  if flag == "Active"
                     else fill_red    if flag == "Disabled"
                     else fill_yellow)
        for ci, key in enumerate(COLS, 1):
            val = r.get(key)
            c   = ws.cell(row=ri, column=ci, value=val)
            c.border = border
            if key == "nach_active":
                c.fill = row_fill
                c.font = Font(bold=True)
            elif key == "emi_amount":
                c.number_format = NUM
            elif key == "nach_enabled":
                c.fill = (fill_green  if val is True
                          else fill_red if val is False
                          else fill_yellow)

    for ci in range(1, len(COLS) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=ci).value or ""))
            for r in range(1, min(len(rows) + 2, 300))
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 38)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Summary tab ───────────────────────────────────────────────────────────
    ws2     = wb.create_sheet("Summary")
    total   = sum(v["count"] for v in active_s.values())
    total_emi = sum(v["emi"] for v in active_s.values())
    act_cnt = active_s.get("Active", {}).get("count", 0)
    act_emi = active_s.get("Active", {}).get("emi", 0)
    cur     = 1

    def section_header(row, text):
        c = ws2.cell(row=row, column=1, value=text)
        c.fill, c.font = sec_fill, sec_font
        c.alignment = Alignment(horizontal="left", indent=1)
        ws2.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=5)
        return row + 1

    def table_header(row, headers):
        for ci, h in enumerate(headers, 1):
            c = ws2.cell(row=row, column=ci, value=h)
            c.fill, c.font = hdr_fill, hdr_font
            c.alignment = Alignment(horizontal="center")
            c.border = border
        return row + 1

    def data_row(row, values, fill, fmt_map=None):
        for ci, val in enumerate(values, 1):
            c = ws2.cell(row=row, column=ci, value=val)
            c.border = border
            c.fill   = fill if not isinstance(fill, list) else fill[ci - 1]
            if fmt_map and (ci - 1) in fmt_map:
                c.number_format = fmt_map[ci - 1]
        return row + 1

    def total_row(row, values, fmt_map=None):
        for ci, val in enumerate(values, 1):
            c = ws2.cell(row=row, column=ci, value=val)
            c.fill = tot_fill
            c.font = Font(bold=True)
            c.border = border
            if fmt_map and (ci - 1) in fmt_map:
                c.number_format = fmt_map[ci - 1]
        return row + 1

    FMTS = {1: NUM, 2: PCT, 3: NUM, 4: PCT}   # col indices 1-4 after label

    # KPI banner
    kpis = [
        ("Month",             month_label),
        ("Total EMIs Due",    f"{total:,}"),
        ("NACH Active",       f"{act_cnt:,}  ({act_cnt/total*100:.1f}%)"),
        ("Active EMI Demand", f"Rs {act_emi:,.0f}"),
        ("Total EMI Demand",  f"Rs {total_emi:,.0f}"),
    ]
    for label, value in kpis:
        lc = ws2.cell(row=cur, column=1, value=label)
        lc.fill, lc.font, lc.border = kpi_fill, kpi_font, border
        vc = ws2.cell(row=cur, column=2, value=value)
        vc.fill, vc.border = kpi_fill, border
        ws2.merge_cells(start_row=cur, start_column=2,
                        end_row=cur, end_column=5)
        cur += 1
    cur += 1

    # Section 1 — NACH Active Status
    cur = section_header(cur, "1.  NACH Active Status  (Payment_setunach.enabled on latest attempt)")
    cur = table_header(cur, ["NACH Status", "Count", "% of Total", "EMI Demand (Rs)", "% of EMI"])
    flag_fills = {"Active": fill_green, "Disabled": fill_red, "Not Yet Attempted": fill_yellow}
    for flag in ["Active", "Disabled", "Not Yet Attempted"]:
        v    = active_s.get(flag, {"count": 0, "emi": 0})
        pct  = round(v["count"] / total       * 100, 1) if total     else 0
        epct = round(v["emi"]   / total_emi   * 100, 1) if total_emi else 0
        cur  = data_row(cur, [flag, v["count"], pct, v["emi"], epct],
                        flag_fills[flag], FMTS)
    cur = total_row(cur, ["TOTAL", total, 100.0, total_emi, 100.0], FMTS)
    cur += 1

    # Section 2 — By Mandate Status
    cur = section_header(cur, "2.  By Mandate Status  (nach_setumandate.status)")
    cur = table_header(cur, ["Mandate Status", "Count", "% of Total", "EMI Demand (Rs)", "% of EMI"])
    m_emi = sum(v["emi"] for v in mandate_s.values())
    for mstatus, v in sorted(mandate_s.items(), key=lambda x: -x[1]["count"]):
        pct  = round(v["count"] / total * 100, 1) if total else 0
        epct = round(v["emi"]   / m_emi * 100, 1) if m_emi else 0
        fill = (fill_green  if mstatus == "live"
                else fill_red if mstatus in ("revoked", "paused")
                else fill_yellow)
        cur  = data_row(cur, [mstatus, v["count"], pct, v["emi"], epct], fill, FMTS)
    cur = total_row(cur, ["TOTAL", total, 100.0, m_emi, 100.0], FMTS)
    cur += 1

    # Section 3 — By Latest Payment Status
    cur = section_header(cur, "3.  By Latest Payment Status  (most recent Payment_setunach.status)")
    cur = table_header(cur, ["Last Payment Status", "Count", "% of Total", "EMI Demand (Rs)", "% of EMI"])
    p_emi = sum(v["emi"] for v in payment_s.values())
    for pstatus, v in sorted(payment_s.items(), key=lambda x: -x[1]["count"]):
        pct  = round(v["count"] / total * 100, 1) if total else 0
        epct = round(v["emi"]   / p_emi * 100, 1) if p_emi else 0
        fill = (fill_green  if pstatus == "success"
                else fill_red if pstatus in ("failed", "no attempt yet")
                else fill_yellow)
        cur  = data_row(cur, [pstatus, v["count"], pct, v["emi"], epct], fill, FMTS)
    cur = total_row(cur, ["TOTAL", total, 100.0, p_emi, 100.0], FMTS)

    ws2.column_dimensions["A"].width = 40
    for col in ["B", "C", "D", "E"]:
        ws2.column_dimensions[col].width = 20

    wb.save(out_path)
    print(f"Excel saved  -> {out_path}")


# ── Chart ─────────────────────────────────────────────────────────────────────
def save_chart(summaries, month_label: str, out_path: str):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker

    active_s = summaries[0]
    labels   = ["Active", "Disabled", "Not Yet Attempted"]
    counts   = [active_s.get(l, {"count": 0})["count"] for l in labels]
    colors   = ["#2ecc71", "#e74c3c", "#f39c12"]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 6))
    fig.suptitle(f"Setu NACH Status - {month_label}", fontsize=14, fontweight="bold")

    bars = ax1.bar(labels, counts, color=colors, width=0.5)
    for bar, cnt in zip(bars, counts):
        if cnt:
            ax1.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 50,
                     f"{cnt:,}", ha="center", va="bottom", fontsize=10, fontweight="bold")
    ax1.set_ylabel("Installments")
    ax1.set_title("Count by NACH Status")
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{int(v):,}"))
    ax1.grid(axis="y", linestyle="--", alpha=0.4)

    non_zero     = [(l, c, col) for l, c, col in zip(labels, counts, colors) if c > 0]
    total        = sum(c for _, c, _ in non_zero)
    wedge_labels = [f"{l}\n{c:,}\n({c/total*100:.1f}%)"
                    for l, c, _ in non_zero]
    ax2.pie([c for _, c, _ in non_zero],
            labels=wedge_labels,
            colors=[col for _, _, col in non_zero],
            startangle=90,
            wedgeprops={"edgecolor": "white", "linewidth": 1.5})
    ax2.set_title("Distribution")

    plt.tight_layout()
    plt.savefig(out_path, dpi=150)
    plt.close()
    print(f"Chart saved  -> {out_path}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    target = (
        datetime.strptime(sys.argv[1], "%Y-%m") if len(sys.argv) > 1
        else datetime(date.today().year, date.today().month, 1)
    )
    year, month = target.year, target.month
    next_m      = date(year + (month // 12), (month % 12) + 1, 1)
    month_start = f"{year}-{month:02d}-01"
    month_end   = next_m.strftime("%Y-%m-%d")
    month_label = target.strftime("%B %Y")
    slug        = f"{year}_{month:02d}"

    print(f"Authenticating to {BASE_URL} ...")
    session, token, csrf = get_auth()
    print(f"Auth OK | Kosh DB | Month: {month_label}")

    print("Fetching Setu NACH data ...")
    rows = fetch_rows(session, token, csrf, month_start, month_end)
    print(f"Fetched {len(rows):,} rows")

    summaries = build_summary(rows)
    print_report(summaries, month_label)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    xlsx_path  = os.path.join(OUTPUT_DIR, f"setu_nach_sheet_{slug}.xlsx")
    chart_path = os.path.join(OUTPUT_DIR, f"setu_nach_chart_{slug}.png")

    print("Writing Excel ...")
    save_excel(rows, summaries, month_label, xlsx_path)
    save_chart(summaries, month_label, chart_path)


if __name__ == "__main__":
    main()
