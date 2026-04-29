"""
AUM & Disbursement MIS — 8 cuts, time-series from 2022-01-01 to latest snapshot
- Each sheet shows all monthly snapshots (1st of each month) from 2022-01-01 onwards
- AUM  : SUM(principal_outstanding) per snapshot
- Disb : SUM(amount * adj_factor) for loans disbursed since FY_FROM, present in that snapshot
         adj_factor: 2026-01 x1.29 | 2026-02 x1.12 | 2026-03 x1.12 | else x1.0
- Snapshot filter : date_created::date >= '2022-01-01' AND EXTRACT(day FROM date_created) = 1
- FY filter       : db_month >= '2022-01'
"""
import os, sys, time, requests
import pandas as pd
import urllib3
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
sys.stdout.reconfigure(encoding='utf-8')

BASE      = "https://superset.bkosh.com"
UN        = os.environ["SUPERSET_UN"]
PW        = os.environ["SUPERSET_PASS"]
DB_ID     = 21
TBL       = 'public."dss_KOSHSUPERSET_all_cohorts_updated_locations_kosh"'
FY_FROM   = '2022-01'
SNAP_FROM = '2022-01-01'

# ── Auth ──────────────────────────────────────────────────────────────────────
session = requests.Session()
session.headers.update({"Referer": f"{BASE}/sqllab/", "User-Agent": "Mozilla/5.0"})

def _do_login():
    for attempt in range(1, 5):
        try:
            r = session.post(f"{BASE}/api/v1/security/login",
                json={"username": UN, "password": PW, "provider": "db", "refresh": True},
                timeout=60)
            tok = r.json()["access_token"]
            cr = session.get(f"{BASE}/api/v1/security/csrf_token/",
                headers={"Authorization": f"Bearer {tok}"}, timeout=30).json()["result"]
            session.headers.update({
                "Authorization": f"Bearer {tok}",
                "X-CSRFToken": cr,
                "Content-Type": "application/json"
            })
            return
        except Exception as e:
            print(f"  [AUTH] Login attempt {attempt} failed: {e} — retrying in 10s")
            time.sleep(10)
    raise RuntimeError("Could not authenticate after 4 attempts")

_do_login()
print(f"[OK] Logged in | Snapshots from: {SNAP_FROM} | FY from: {FY_FROM}")

def refresh_auth():
    _do_login()
    print("  [AUTH] Re-authenticated")

def run(sql, label="", _retry=True):
    try:
        r = session.post(f"{BASE}/api/v1/sqllab/execute/",
            json={"database_id": DB_ID, "sql": sql, "json": True, "queryLimit": 20000},
            timeout=300)
    except Exception as e:
        print(f"  [ERR] {label}: request failed — {e}")
        return pd.DataFrame()

    if not r.content:
        print(f"  [ERR] {label}: empty response (server timeout)")
        return pd.DataFrame()

    if r.status_code == 404 and _retry:
        refresh_auth()
        return run(sql, label, _retry=False)

    try:
        d = r.json()
    except Exception:
        print(f"  [ERR] {label}: bad JSON — HTTP {r.status_code} — {r.text[:200]}")
        return pd.DataFrame()

    errs = d.get("error") or d.get("errors")
    if errs:
        print(f"  [ERR] {label}: {str(errs)[:200]}")
        refresh_auth()
        return pd.DataFrame()

    rows = d.get("data", [])
    if not rows:
        print(f"  [WARN] No data: {label}")
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    print(f"  [OK] {label} -- {len(df)} rows across {df['snapshot_date'].nunique() if 'snapshot_date' in df.columns else '?'} snapshots")
    return df


YEARS = [2022, 2023, 2024, 2025, 2026]

def run_batched(sql_template, label):
    parts = []
    for yr in YEARS:
        year_sql = sql_template.replace(
            "{year_filter}",
            f"AND date_created::date >= '{yr}-01-01' AND date_created::date < '{yr+1}-01-01'"
        )
        chunk = run(year_sql, f"{label} ({yr})")
        if not chunk.empty:
            parts.append(chunk)
    if not parts:
        return pd.DataFrame()
    df = pd.concat(parts, ignore_index=True)
    print(f"  [COMBINED] {label} -- {len(df)} total rows")
    return df

# ── Shared filter & expressions ───────────────────────────────────────────────
AUM_WHERE = f"""
    date_created::date >= '{SNAP_FROM}'
    AND EXTRACT(day FROM date_created) = 1
    AND amount > 0
"""

ADJ = """amount::numeric * CASE
        WHEN db_month = '2026-01' THEN 1.29
        WHEN db_month = '2026-02' THEN 1.12
        WHEN db_month = '2026-03' THEN 1.12
        ELSE 1.0
    END"""

FY_CASE = f"CASE WHEN db_month >= '{FY_FROM}' THEN {ADJ} ELSE 0 END"
FY_CNT  = f"CASE WHEN db_month >= '{FY_FROM}' THEN loanshare_id ELSE NULL END"

# ── Employer Sector mapping ────────────────────────────────────────────────────
EW_TBL = 'public."dss_KOSHSUPERSET_employer_wise_"'

_SECTORS = [
    ("Auto / Automotive", [
        # OEMs & major brands
        "maruti","honda car","honda motorcycle","honda logistics","honda india",
        "bajaj auto","hero moto","tvs motor","tvs sundram","tvs srichakra",
        "mahindra","tata motor","tata autocomp","yamaha motor","ford india",
        "hyundai india","renault india","skoda india","volkswagen india",
        # Tier-1 / Tier-2 suppliers
        "varroc","uno minda","minda limited","minda corporation","sandhar",
        "sansera","rockman","badve","uniparts","lumax auto","denso haryana",
        "escorts kubota","l.g. balakrishnan","metalman","satyam auto","pp auto",
        "ampp auto","victoria auto","victora auto","afflatus","spark minda",
        "vega auto","shivam auto","fcc clutch","fcc india","fcc limited",
        "jk tyre","imperial auto","neel auto","happy forgings","qh talbros",
        "alf engineering","motherson sumi","motherson","wheels india",
        "lan engineering","endurance technologies","endurance sector",
        "endurance limited","endurance pvt","endurance india","endurance ",
        "polar auto","sunbeam auto","makino auto","napino","narvada","belrise",
        "bajajsons","singla forging","hindustan tyre","a. k. automatics",
        "vgr engineering","fcc auto","elite engineering","universal autofoundry",
        "ksec pvt","micro turners","k.s.e.c","ask automotive","ask auto",
        "sarita honda","mmt micro","krishna maruti","fcc",
        # Generic auto patterns
        "auto pvt ltd","automotive pvt","motors pvt ltd","motor pvt ltd",
        "auto components pvt","auto industries pvt","auto tech pvt",
        "auto engineering pvt","auto parts pvt","vehicle pvt",
    ]),
    ("Pharma / Healthcare", [
        # Known companies
        "akums drugs","akums pharmaceut","akums pharma","multani pharma",
        "multani pharmaceut","synokem","romsons","malik lifescience","axa parenteral",
        "indian herbals","om sai pharma","ankura hospital","tmu hospital",
        "nanz pharma","cipla","sun pharma","dr reddy","lupin","mankind pharma",
        "glenmark","torrent pharma","ipca","alkem","zydus","sun pharmaceutical",
        "abbott india","pfizer india","sanofi india","novartis india",
        "apollo hospital","fortis hospital","max hospital","medanta",
        "medplus","healthkart","med pvt","pharma pvt",
        # Generic patterns
        "hospital","pharma","pharmaceutical","healthcare","medikit","medic",
        "clinic","nursing home","dispensary","life science","diagnostic",
        "pathology","medical college","nursing pvt","health pvt",
    ]),
    ("Textiles / Apparel", [
        # Known companies
        "richa global","richaco","relaxo footwear","relaxo foot","relaxo limited",
        "lakhani footwear","lakhani foot","lamba footwear","liberty shoes",
        "liberty shoe","sangam india","sangam spinner","sangam group","modelama",
        "sarita handa","devgiri","spring overseas","bindu fashion","basant overseas",
        "pinnacle clothing","anubhav apparel","raj overseas","tangerine skies",
        "vision export","ir export","aman export","kaleen","rajlakshmi cotton",
        "golden texo","designco export","rswm","knitwell","c.l. gupta overseas",
        "cl gupta overseas","c l gupta overseas","c.l gupta overseas",
        "rajasthan textile","rajasthan textile mills","birla textile","sutlej textile",
        "bhadohi weavetax","campus activewear","gupta h.c.","mochiko shoe","mochiko",
        "cta apparel","the shivalika","ratan textile","nahar","pearl precision",
        "fa home and app","paramount products","virola international","abros sports",
        "c.l gupta export","cl gupta export","c l gupta export","ginni international",
        "lakhani","sarita handa export","krishna textile","indian sport pvt",
        "trela footwear","ratan textiles","bco impex",
        # Generic patterns
        "textile pvt","textil","apparel pvt","apparel","garment pvt","garment",
        "weaving pvt","weaving","knitwear","hosiery pvt","hosiery",
        "dyeing pvt","footwear pvt","carpet pvt","spinning pvt",
        "fabric pvt","woven pvt","yarn pvt","embroidery pvt",
    ]),
    ("Consumer Goods / FMCG", [
        # Known companies
        "patanjali","nestle india","hindustan unilever","hul pvt","emami",
        "kent ro","britannia","parle agro","parle pvt","milton company","milton pvt",
        "b.l agro","hamilton housewares","century pulp","mahesh edible",
        "lifelong india","voltas","dabur","reckitt","rackitt","vdsd foods",
        "dolphin sector 2","itc limited","perfetti van melle","titan company",
        "marico","godrej consumer","procter & gamble","p&g","colgate","gillette",
        "havmore","amul","mother dairy","kwality","lotus beauty","lotus herbal",
        "hindustan lever","lever pvt","bajaj consumer","bajaj almond",
        "himalaya drug","himalaya pvt","nirma pvt","ghadi detergent",
        "ghadi pvt","vi john","vi-john","joy pvt","parachute pvt",
        "vicco pvt","boroline pvt","fair & lovely",
        # Generic
        "fmcg","consumer goods","household goods","personal care pvt",
    ]),
    ("Rubber / Industrial", [
        "designco pvt","a.g. industries pvt","pyoginam private","bkt tyre",
        "balkrishna industries","century plyboard","classic industries pvt",
        "geeken seating","kanchan india","dvs industries","shree amba industries",
        "pritam international pvt","laopla","m r enterprises","mantri mettalic",
        "mppl pvt","jain cord","omega printopack","luxor writing","cello housware",
        "pg electroplast","karam udhyog","govind rubber","fcc pvt",
        "itc company","rubber pvt","tyre pvt","plyboard pvt","foam pvt",
        "polyurethane","gasket pvt","seal pvt","plastic pvt","polymer pvt",
    ]),
    ("Electronics / Technology", [
        # Known companies
        "studds accessories","fiem industries","anchor by panasonic","anchor panasonic",
        "vvdn technolog","epak durable","delta power solution","surya roshni",
        "luminous power","c&s electric","c & s electric","polycab","time technoplast",
        "myra techno","bmr pvt","apj investment pvt","tej shoe tech","okaya",
        "laxmi remote","jns instrument","global medikit","lifelong meditech",
        "wonder electr","r k lighting","wipro","c&s electricals","c & s electricals",
        "samsung india","samsung pvt","samsung electronics","siemens india","abb india",
        "schneider electric","legrand","tata elxsi","infosys","hcl tech",
        "tech mahindra","accenture","capgemini","cognizant","mphasis","zensar",
        "l&t technology","honeywell","philips india","orient electric","bajaj electric",
        "crompton greaves","finolex cable","vaibhav global","akal information",
        "noetech","vguard","v guard","v-guard","samriddhi automation",
        "samriddhi automations","bosch india",
        # Generic
        "electronic pvt","electron pvt","technolog pvt","software pvt",
        "it pvt","digital pvt","it service","automation pvt",
    ]),
    ("Manufacturing / Industrial", [
        "havells india","havells company","havells limited","amber enterprises",
        "amber pvt","amber ltd","dixon technolog","meenakshi polymer",
        "asahi india glass","goldplus glass","globe hi tech","man machine solutions",
        "man machines solutions","man machines solution","manmachine management",
        "aurangabad electrical","v guard industry","neel metal products ltd haridwar",
        "neel metal ltd haridwar","flexituff","blue heavens","frontier alloy",
        "johnson controls","greaves cotton","thermax","bharat forge","isgec",
        "kunstocom","hnv casting","wonder electricals limited",
        "wonder electrical limited","r k lighting pvt","c&s electricals limited",
        "panchkula steel","super alloy casting","padmavati pipes",
        "universal autofoundry","rp associate","rp associates",
        # Generic industrial patterns
        "industries pvt ltd","casting pvt","forging pvt","engineering pvt ltd",
        "fab pvt","alloy pvt","foundry pvt","fabrication pvt",
        "tooling pvt","moulding pvt","precision pvt",
    ]),
    ("Manufacturing / Metal Products", [
        "hnv casting india","panchkula steel pvt","super alloy castings",
        "padmavati pipes & fitting","steel pvt ltd","metal products pvt",
        "alloys pvt ltd","stainless steel pvt","iron pvt","copper pvt",
        "aluminium pvt","metal casting pvt","sheet metal pvt",
        "wire pvt","pipe pvt","fitting pvt","steel india pvt","casting india pvt",
    ]),
    ("Manufacturing / Electrical", [
        "aurora electric","surya electrical","rk electrical",
        "electrical pvt ltd","switchgear pvt","transformer pvt",
        "pump pvt","generator pvt","cable pvt",
    ]),
    ("Manufacturing / Rubber & Tyres", [
        "govind rubber ltd","govind rubber","apollo tyre","ceat tyre","mrf tyre",
        "birla tyre","balkrishna tyre","continental pvt","michelin india",
        "rubber pvt ltd","tyre pvt ltd",
    ]),
    ("Food Processing", [
        "roquette india","krown agro","greendot health foods","avitech nutrition",
        "advance panels","itc agri","dabur agri","britannia industries",
        "parle products","nestle food","agro pvt","food processing pvt",
        "agro food pvt","food pvt ltd","bakery pvt","flour mill","sugar mill",
        "rice mill","dal mill","spice pvt","confection pvt","snack pvt",
        "beverage pvt","jam pvt","pickle pvt","dairy pvt",
    ]),
    ("Real Estate / Construction", [
        "unimax international","balaji action buildwell","avadh rail infra",
        "genus power infra","kirby building","ultratech cement","dlf limited",
        "l&t construction","larsen & toubro","tata projects","shapoorji",
        "sobha","prestige estate","brigade group","godrej properties",
        "mmt distributors pvt","mmt distributors",
        "cement pvt","construction pvt","buildwell","infra pvt","real estate",
        "housing pvt","builder pvt","realty pvt","township pvt",
    ]),
    ("Logistics / Transport", [
        "zinka logistics","loadshare networks","dhl india","bluedart","fedex india",
        "ekart logistics","xpressbees","shadowfax","ecom express","delhivery",
        "safexpress","ezidrive","tour and travel","travels pvt","roadways pvt",
        "road transport pvt","cab service pvt","taxi pvt","fleet pvt",
        "transport pvt","logistics pvt","cargo pvt","trucking pvt",
        "courier pvt","supply chain pvt","shipping pvt","freight pvt",
        "warehousing pvt","sis limited",
    ]),
    ("Services / Security", [
        "sis security","proactive security","gdx security","checkmate security",
        "goldeneye guarding","ashoka unique services","securitas","g4s pvt",
        "security services pvt","security solutions pvt","security guard pvt",
        "guarding pvt","surveillance pvt","detective pvt","patrol pvt",
    ]),
    ("Services / Staffing", [
        "sk services","altum staffing","manpower pvt","quess corp",
        "kapston facilities","arcos skill","teamlease","mafoi","adecco","randstad",
        "staffing pvt","placement pvt","staffing services","workforce pvt",
        "human resource pvt","hr service pvt","recruitment pvt",
        "flexi staffing","manpower services","outsourcing pvt",
    ]),
    ("Services / Facilities Management", [
        "nimbus harbor","yes madam","urbanclap","urban company",
        "housekeeping pvt","catering service pvt","facility management pvt",
        "facilities management pvt","integrated facility",
        "maintenance service pvt","pest control pvt","cleaning service pvt",
    ]),
    ("Hospitality", [
        "hotel","resort","motel","lodge pvt","inn pvt","restaurant pvt",
        "dessertino","dhaba pvt","banquet pvt","cafe pvt","canteen pvt",
        "hospitality pvt","food court pvt","club pvt","spa pvt",
    ]),
    ("Government / PSU", [
        "sjvn limited","sjvn","cpwd","ntpc limited","ntpc ltd","ongc",
        "hpcl","bpcl","iocl","coal india","power grid","nhpc","bhel",
        "bsnl","air india","indian railways","nagarpalika","municipal corporation",
        "panchayat","government","govt ","public sector","police department",
        "army","navy","air force","central govt","state govt",
        "pwd ","nhai","gail india","sail ","hpcl mittal","postal dept","lic india",
    ]),
    ("BFSI / Finance", [
        "finnew corporate","paytm","hdfc bank","icici bank","sbi ","axis bank",
        "kotak bank","yes bank","idfc bank","bandhan bank","bajaj finserv",
        "bajaj finance","mahindra finance","shriram finance","muthoot",
        "manappuram","fullerton","capital first","piramal finance","aequitas",
        "loantap","cashe","nira finance","kreditbee","moneyview","navi pvt",
        "aditya birla capital","tata capital","insurance pvt","nbfc pvt",
        "microfinance pvt","credit pvt","lending pvt","financial service pvt",
        "wealth management pvt","securities pvt","broker pvt",
    ]),
    ("Chemicals / Specialty", [
        "srf ltd","macnorr mcnore","india pesticides","atul chemicals",
        "kansai nerolac","kansai nerlac","asian paints","berger paints",
        "nippon paint","akzo nobel","jk cement","acc cement","ambuja cement",
        "aarti industries","vinati organics","specialty chemical",
        "chemical pvt ltd","chemicals pvt ltd","petrochemical pvt",
        "lubricant pvt","adhesive pvt","coating pvt","paint pvt","ink pvt","resin pvt",
    ]),
    ("Fertilizers / Chemicals", [
        "chambal fertilizers","chambal fertiliser","coromandel international",
        "deepak fertiliser","national fertilizer","iffco","kribhco",
        "fertilizer pvt","fertiliser pvt","agrochem pvt",
    ]),
    ("Glass / Materials", [
        "greenlam industries","greenpanel industries","aica laminates",
        "neelmatel","asahi india glass","saint gobain","pilkington",
        "hindalco","vedanta","laminate pvt","glass pvt","marble pvt",
        "granite pvt","tile pvt","ceramic pvt","flooring pvt",
    ]),
    ("Printing / Packaging", [
        "parksons packaging","thomson press india","flexituff ventures",
        "manjushree technopack","printing pvt","packaging pvt","carton pvt",
        "paper pvt","corrugated pvt","label pvt","polybag pvt","sachet pvt",
    ]),
    ("Exports / Trading", [
        "abc impex","ac brothers","a.c. brothers","gaurav international",
        "san international","sahu global","evergreen international",
        "a.c brothers","raj overseas corporate","import export pvt",
        "international trading","export pvt ltd","import pvt",
        "trading pvt ltd","trading co pvt","mercantile pvt","merchandise pvt",
    ]),
    ("Telecom / Media", [
        "airtel","jio pvt","vodafone","tata teleservices","mtnl",
        "dish tv","tata sky","sun direct","zee media","tv18","network18",
        "telecom pvt","media pvt","broadcasting pvt","content pvt",
    ]),
    ("Toys / Misc Manufacturing", [
        "neel mattel","mattel india","funskool","hasbro india","toy pvt",
        "games pvt","sporting goods pvt","sports equip pvt","indian sport pvt",
        "abros sports",
    ]),
    ("Education", [
        "tmu university","tmu college","harrow senior secondary",
        "prerna engineering education","school","college","university",
        "vidyalaya","academy pvt","coaching pvt","education pvt",
        "institute pvt","tutorial pvt","learning pvt",
    ]),
    ("Self-employed / Tailoring", [
        "tailor","silai mashin","silai machine","sewing work","stitching work",
        "darzi","privet job tailor","tailoring work","silai kaam","kapda silai",
    ]),
    ("Self-employed / Dairy", [
        "milk dairy","milk deiry","dairy farm","milk and dairy","dudh ka kam",
        "private job milk","dairy product","milk supply","milkman","doodh",
    ]),
]

def _build_sector_case():
    lines = ["CASE"]
    for sector, keywords in _SECTORS:
        conds = " OR ".join(f"ew.employer ILIKE '%{kw}%'" for kw in keywords)
        lines.append(f"    WHEN ew.employer IS NOT NULL AND ({conds})")
        lines.append(f"        THEN '{sector}'")
    lines.append("    WHEN ew.employer IS NULL OR TRIM(ew.employer) = '' THEN 'Unknown'")
    lines.append("    ELSE 'Other / Unclassified'")
    lines.append("END")
    return "\n".join(lines)

SECTOR_CASE = _build_sector_case()

# ── Sanity check ──────────────────────────────────────────────────────────────
print("\n[Sanity Check — available snapshots]")
run(f"""
SELECT
    date_created::date           AS snapshot_date,
    COUNT(DISTINCT loanshare_id) AS active_loans,
    SUM(principal_outstanding)::bigint AS total_aum
FROM {TBL}
WHERE {AUM_WHERE}
GROUP BY 1
ORDER BY 1
""", "Snapshots")

# ── Template ──────────────────────────────────────────────────────────────────
def cut(dim_expr, label, order_by='"AUM" DESC'):
    tmpl = f"""
SELECT
    date_created::date             AS snapshot_date,
    {dim_expr}                     AS dimension,
    SUM(principal_outstanding)::bigint AS "AUM",
    SUM({FY_CASE})::bigint         AS disbursal
FROM {TBL}
WHERE {AUM_WHERE}
  {{year_filter}}
GROUP BY 1, 2
ORDER BY snapshot_date, {order_by}
"""
    return run_batched(tmpl, label)

# ── 8 cuts ────────────────────────────────────────────────────────────────────
print("\n[1] By Cluster")
df1 = cut("COALESCE(cluster_name, 'Unknown')", "1. By Cluster")

print("\n[2] By Ticket Size")
df2 = run_batched(f"""
SELECT
    date_created::date AS snapshot_date,
    CASE
        WHEN amount <= 25000  THEN '1. Upto 25K'
        WHEN amount <= 50000  THEN '2. 25K - 50K'
        WHEN amount <= 100000 THEN '3. 50K - 1L'
        WHEN amount <= 200000 THEN '4. 1L - 2L'
        WHEN amount <= 500000 THEN '5. 2L - 5L'
        ELSE                       '6. Above 5L'
    END                    AS dimension,
    SUM(principal_outstanding)::bigint AS "AUM",
    SUM({FY_CASE})::bigint AS disbursal
FROM {TBL}
WHERE {AUM_WHERE}
  {{year_filter}}
GROUP BY 1, 2
ORDER BY snapshot_date, dimension
""", "2. By Ticket Size")

print("\n[3] By Sourcing Mix")
df3 = cut("""CASE sales_channel
        WHEN 'fso' THEN 'Direct (Field Sales)'
        WHEN 'ref' THEN 'DSA / Referral'
        WHEN 'lsp' THEN 'Partnerships (LSP)'
        ELSE COALESCE(sales_channel, 'Unknown')
    END""", "3. By Sourcing Mix")

print("\n[4] By Income Level")
df4 = run_batched(f"""
SELECT
    date_created::date AS snapshot_date,
    CASE
        WHEN annual_salary IS NULL    THEN '0. Not Available'
        WHEN annual_salary <= 120000  THEN '1. Upto 10K/month'
        WHEN annual_salary <= 240000  THEN '2. 10K - 20K/month'
        WHEN annual_salary <= 360000  THEN '3. 20K - 30K/month'
        WHEN annual_salary <= 600000  THEN '4. 30K - 50K/month'
        ELSE                               '5. Above 50K/month'
    END                    AS dimension,
    SUM(principal_outstanding)::bigint AS "AUM",
    SUM({FY_CASE})::bigint AS disbursal
FROM {TBL}
WHERE {AUM_WHERE}
  {{year_filter}}
GROUP BY 1, 2
ORDER BY snapshot_date, dimension
""", "4. By Income Level")

print("\n[5] By Employer Sector")
df5 = run_batched(f"""
SELECT
    c.date_created::date                  AS snapshot_date,
    {SECTOR_CASE}                         AS dimension,
    SUM(c.principal_outstanding)::bigint  AS "AUM",
    SUM({FY_CASE})::bigint                AS disbursal
FROM {TBL} c
LEFT JOIN {EW_TBL} ew ON ew.loanshare_id = c.loanshare_id
WHERE {AUM_WHERE}
  {{year_filter}}
GROUP BY 1, 2
ORDER BY snapshot_date, "AUM" DESC
""", "5. By Employer Sector")

print("\n[6] By Lender Count")
df6 = run_batched(f"""
SELECT
    date_created::date AS snapshot_date,
    CASE
        WHEN ranking IS NULL OR ranking::int = 1 THEN '1. Exclusive to Kosh'
        WHEN ranking::int = 2                    THEN '2. With 1 other lender'
        WHEN ranking::int = 3                    THEN '3. With 2 other lenders'
        ELSE                                          '4. With >2 lenders'
    END                    AS dimension,
    SUM(principal_outstanding)::bigint AS "AUM",
    SUM({FY_CASE})::bigint AS disbursal
FROM {TBL}
WHERE {AUM_WHERE}
  {{year_filter}}
GROUP BY 1, 2
ORDER BY snapshot_date, dimension
""", "6. By Lender Count")

print("\n[7] By Book Type")
df7 = cut("""CASE
        WHEN lender IN ('light_nfcpl_colending','light_colending',
                        'light_colending_new','hindon_colending',
                        'janasha_colending','fintree_colending') THEN 'Co-lending'
        WHEN lender = 'narendra_finance'                         THEN 'BC Model'
        WHEN lender IS NULL                                      THEN 'Own Book'
        ELSE lender
    END""", "7. By Book Type")

print("\n[8] New vs Repeat")
df8 = run_batched(f"""
SELECT
    date_created::date AS snapshot_date,
    CASE cx_type
        WHEN '1_time_cx'    THEN 'New (1st loan)'
        WHEN 'recurring_cx' THEN 'Repeat (2nd+ loan)'
        ELSE COALESCE(cx_type, 'Unknown')
    END                    AS dimension,
    COUNT(DISTINCT loanshare_id) AS active_borrowers
FROM {TBL}
WHERE {AUM_WHERE}
  {{year_filter}}
GROUP BY 1, 2
ORDER BY snapshot_date, dimension
""", "8. New vs Repeat")

datasets = {
    "1. By Cluster":         df1,
    "2. By Ticket Size":     df2,
    "3. By Sourcing Mix":    df3,
    "4. By Income Level":    df4,
    "5. By Employer Sector":  df5,
    "6. By Lender Count":    df6,
    "7. By Book Type":       df7,
    "8. New vs Repeat":      df8,
}

# ── Styles ────────────────────────────────────────────────────────────────────
from openpyxl import Workbook

TITLE_FILL = PatternFill("solid", fgColor="17324D")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=11)
DATE_FILL  = PatternFill("solid", fgColor="1F4E79")
DATE_FONT  = Font(color="FFFFFF", bold=True, size=8)
DIM_FILL   = PatternFill("solid", fgColor="D6E4F0")
DIM_FONT   = Font(bold=True, size=9, color="17324D")
EVEN_FILL  = PatternFill("solid", fgColor="FFFFFF")
ODD_FILL   = PatternFill("solid", fgColor="EBF3FB")
TOTAL_FILL = PatternFill("solid", fgColor="17324D")
TOTAL_FONT = Font(color="FFFFFF", bold=True, size=9)
BORDER     = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)
MONEY_FMT = "#,##0"

def _border_cell(cell, fill, font=None, align=None, fmt=None):
    cell.fill   = fill
    cell.border = BORDER
    if font:  cell.font           = font
    if align: cell.alignment      = align
    if fmt:   cell.number_format  = fmt

def write_pivot_block(ws, pivot, start_row, label, fmt=MONEY_FMT):
    n_dims, n_cols = pivot.shape
    hdr_row        = start_row + 1
    first_data_row = start_row + 2
    last_data_row  = start_row + 1 + n_dims
    total_row      = last_data_row + 1

    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row,   end_column=n_cols + 1)
    c = ws.cell(row=start_row, column=1, value=label)
    c.fill      = TITLE_FILL
    c.font      = TITLE_FONT
    c.border    = BORDER
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[start_row].height = 22

    ws.row_dimensions[hdr_row].height = 64
    corner = ws.cell(row=hdr_row, column=1, value="")
    _border_cell(corner, DATE_FILL)
    for i, d in enumerate(pivot.columns, start=2):
        c = ws.cell(row=hdr_row, column=i, value=d)
        _border_cell(c, DATE_FILL, font=DATE_FONT,
                     align=Alignment(horizontal="center", vertical="bottom",
                                     text_rotation=90, wrap_text=False))

    for r_off, (dim, row_data) in enumerate(pivot.iterrows()):
        row      = first_data_row + r_off
        row_fill = ODD_FILL if r_off % 2 == 1 else EVEN_FILL
        ws.row_dimensions[row].height = 16

        dim_c = ws.cell(row=row, column=1, value=dim)
        _border_cell(dim_c, DIM_FILL, font=DIM_FONT,
                     align=Alignment(horizontal="left", vertical="center", indent=1))

        for c_off, val in enumerate(row_data, start=2):
            v  = int(val) if pd.notna(val) else 0
            dc = ws.cell(row=row, column=c_off, value=v)
            _border_cell(dc, row_fill,
                         align=Alignment(horizontal="right", vertical="center"),
                         fmt=fmt)

    # Total row
    ws.row_dimensions[total_row].height = 16
    tc = ws.cell(row=total_row, column=1, value="Total")
    _border_cell(tc, TOTAL_FILL, font=TOTAL_FONT,
                 align=Alignment(horizontal="left", vertical="center", indent=1))
    col_totals = pivot.fillna(0).sum(axis=0)
    for c_off, val in enumerate(col_totals, start=2):
        dc = ws.cell(row=total_row, column=c_off, value=int(val))
        _border_cell(dc, TOTAL_FILL, font=TOTAL_FONT,
                     align=Alignment(horizontal="right", vertical="center"),
                     fmt=fmt)

    ws.column_dimensions["A"].width = 24
    for i in range(2, n_cols + 2):
        ws.column_dimensions[get_column_letter(i)].width = 8

    return total_row, hdr_row, first_data_row, n_cols


# ── Build workbook ────────────────────────────────────────────────────────────
OUT_FILE = f"AUM_MIS_TimeSeries_{datetime.today().strftime('%Y%m%d_%H%M')}.xlsx"
wb = Workbook()
wb.remove(wb.active)

for name, df in datasets.items():
    ws = wb.create_sheet(title=name[:31])
    ws.sheet_view.showGridLines = False

    if df.empty:
        ws.cell(row=1, column=1, value="No data")
        continue

    is_new_repeat = "active_borrowers" in df.columns
    df = df.copy()
    df["snapshot_date"] = pd.to_datetime(df["snapshot_date"])

    if is_new_repeat:
        piv = df.pivot_table(index="dimension", columns="snapshot_date",
                             values="active_borrowers", aggfunc="sum")
        piv = piv.sort_index(axis=1)
        piv.columns = [c.strftime("%b %Y") for c in piv.columns]
        piv = piv.reindex(piv.iloc[:, -1].sort_values(ascending=False).index)

        last, hdr, fdr, nc = write_pivot_block(ws, piv, start_row=1,
                                                label="Active Borrowers", fmt="#,##0")
        ws.freeze_panes = "B3"

    else:
        aum_piv  = df.pivot_table(index="dimension", columns="snapshot_date",
                                  values="AUM", aggfunc="sum")
        disb_piv = df.pivot_table(index="dimension", columns="snapshot_date",
                                  values="disbursal", aggfunc="sum")
        aum_piv  = aum_piv.sort_index(axis=1)
        disb_piv = disb_piv.sort_index(axis=1)
        date_labels      = [c.strftime("%b %Y") for c in aum_piv.columns]
        aum_piv.columns  = date_labels
        disb_piv.columns = date_labels

        row_order = aum_piv.iloc[:, -1].sort_values(ascending=False).index
        aum_piv  = aum_piv.reindex(row_order)
        disb_piv = disb_piv.reindex(row_order)

        aum_last, aum_hdr, aum_fdr, _ = write_pivot_block(
            ws, aum_piv, start_row=1,
            label="AUM — Principal Outstanding (INR)")

        disb_start = aum_last + 3
        disb_last, disb_hdr, disb_fdr, _ = write_pivot_block(
            ws, disb_piv, start_row=disb_start,
            label="Total Disbursal (INR)")

        ws.freeze_panes = "B3"

wb.save(OUT_FILE)
print(f"\n[DONE] Saved: {OUT_FILE}")
print(f"  Sheets: {', '.join(datasets.keys())}")
